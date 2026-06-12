import json
import re
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = ROOT / "config" / "competitors.yaml"
APRIMORA_MATRIX = Path(r"C:\Users\USUÁRIO\Downloads\Matriz  Análise Competitiva - Aprimora 2.xlsx")
EDUCACIONAL_MATRIX = Path(r"C:\Users\USUÁRIO\Downloads\Análise Concorrentes - Educacional.xlsx")


CANONICAL_NAMES = {
    "Arduíno": "Arduino",
    "Clever US": "Clever",
    "Layers": "Layers Education",
    "Nav a vela": "Nave à Vela",
    "Via Maker": "Viamaker",
    "Wonde UK": "Wonde",
    "Zoom": "ZOOM Education",
    "Max.Ia": "Max.ia",
    "Santana": "Santana (LEGO Education Revenda)",
}

SKIP_NAMES = {"Aprimora", "Inventura", "Micro:bit", "Lego Eeducation"}


def clean(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def split_site(value):
    text = clean(value)
    if not text:
        return ""
    match = re.search(r"https?://[^\s]+", text)
    if match:
        return match.group(0).rstrip(".,)")
    return text.rstrip(".,)")


def slug(value):
    import unicodedata

    text = unicodedata.normalize("NFD", value.lower())
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def flow(items):
    return "[" + ", ".join(json.dumps(item, ensure_ascii=False) for item in items) + "]"


def quote(value):
    return json.dumps(value, ensure_ascii=False)


def parse_flow_list(value):
    if "[" not in value:
        return []
    raw = value[value.index("[") :].strip()
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        return [item.strip().strip('"') for item in raw.strip("[]").split(",") if item.strip()]


def read_current_competitors():
    competitors = []
    current = None
    for line in CONFIG_PATH.read_text(encoding="utf-8").splitlines():
        stripped = line.strip()
        if stripped.startswith("- name:") and line.startswith("  - "):
            if current:
                competitors.append(current)
            current = {
                "name": stripped.split(":", 1)[1].strip().strip('"'),
                "website": "",
                "scope": "competes_market",
                "regions": ["BR"],
                "markets": ["private"],
                "tags": [],
                "products_impacted": [],
            }
        elif current and ":" in stripped:
            key, value = stripped.split(":", 1)
            value = value.strip()
            if key in {"website", "scope"}:
                current[key] = value.strip('"')
            elif key in {"regions", "markets", "tags", "products_impacted", "impacted_products"}:
                target = "products_impacted" if key == "impacted_products" else key
                current[target] = parse_flow_list(value)
    if current:
        competitors.append(current)
    return competitors


def infer_regions(*parts):
    text = " ".join(clean(part).lower() for part in parts)
    if any(term in text for term in ["brasil", "baasil", ".br", "brasileira", "nacional"]):
        return ["BR"]
    if any(term in text for term in ["latam", "américa latina", "america latina"]):
        return ["LATAM"]
    return ["GLOBAL"]


def infer_products(label, name=""):
    text = f"{clean(label)} {clean(name)}".lower()
    products = []
    if "aprimora" in text:
        products.append("Aprimora")
    if "hub" in text or "suíte" in text or "suite" in text:
        products.append("HUB.Educacional")
    if "pense" in text:
        products.append("Pense +")
    if "innovation" in text or "inventura" in text:
        products.append("Inventura")
    if "micro" in text:
        products.append("micro:bit")
    if "lego" in text:
        products.extend(["LEGO Education", "Robotis"])
    if "atto" in text or "robótica" in text or "robotica" in text:
        products.append("Robotis")
    if "mesa" in text or "mesinha" in text:
        products.append("Mesa Educacional")
    return dedupe(products or ["HUB.Educacional"])


def infer_tags(*parts):
    text = " ".join(clean(part).lower() for part in parts)
    tags = ["source:xlsx"]
    rules = [
        ("matemática", "math"),
        ("matematica", "math"),
        ("português", "portuguese"),
        ("portugues", "portuguese"),
        ("língua portuguesa", "portuguese"),
        ("leitura", "reading"),
        ("redação", "writing"),
        ("texto", "writing"),
        ("game", "gamification"),
        ("gamifica", "gamification"),
        ("adaptativa", "adaptive learning"),
        ("personalizado", "personalized learning"),
        ("avalia", "assessment"),
        ("diagnóstico", "assessment"),
        ("diagnostico", "assessment"),
        ("analytics", "analytics"),
        ("relatório", "analytics"),
        ("dados", "analytics"),
        ("inteligência artificial", "AI"),
        (" ia ", "AI"),
        ("maker", "maker"),
        ("steam", "STEAM"),
        ("stem", "STEM"),
        ("robótica", "robotics"),
        ("robotica", "robotics"),
        ("programação", "coding"),
        ("programacao", "coding"),
        ("micro:bit", "micro:bit"),
        ("lego", "LEGO Education"),
        ("hardware", "hardware"),
        ("kit", "hardware"),
        ("lms", "LMS"),
        ("gestão", "school management"),
        ("gestao", "school management"),
        ("integra", "integration"),
        ("plataforma", "platform"),
        ("formação", "teacher training"),
        ("formacao", "teacher training"),
        ("bncc", "BNCC"),
        ("offline", "offline"),
        ("saas", "SaaS"),
        ("gratuito", "free"),
    ]
    for needle, tag in rules:
        if needle in text:
            tags.append(tag)
    return dedupe(tags)


def dedupe(items):
    result = []
    seen = set()
    for item in items:
        item = clean(item)
        if item and item.lower() not in seen:
            seen.add(item.lower())
            result.append(item)
    return result


def upsert(target, item):
    name = CANONICAL_NAMES.get(item["name"], item["name"])
    if name in SKIP_NAMES:
        return
    item["name"] = name
    key = slug(name)
    existing = target.get(key)
    if not existing:
        target[key] = item
        return

    if not existing.get("website") and item.get("website"):
        existing["website"] = item["website"]
    existing["regions"] = dedupe(existing.get("regions", []) + item.get("regions", []))
    existing["markets"] = dedupe(existing.get("markets", []) + item.get("markets", []))
    existing["tags"] = dedupe(existing.get("tags", []) + item.get("tags", []))
    existing["products_impacted"] = dedupe(existing.get("products_impacted", []) + item.get("products_impacted", []))
    if existing.get("scope") != "competes_market":
        existing["scope"] = item.get("scope", existing["scope"])


def load_aprimora_matrix():
    workbook = openpyxl.load_workbook(APRIMORA_MATRIX, data_only=True, read_only=True)
    links = {}
    sheet = workbook["Planilha1"]
    for col in range(1, sheet.max_column + 1):
        name = clean(sheet.cell(2, col).value)
        site = split_site(sheet.cell(5, col).value)
        if name and site:
            links[name] = site

    summary = workbook["Resumo"]
    headers = [clean(summary.cell(1, col).value) or str(col) for col in range(1, 13)]
    for row in range(2, summary.max_row + 1):
        values = {headers[col - 1]: summary.cell(row, col).value for col in range(1, 13)}
        name = clean(values.get("1"))
        if not name or name in SKIP_NAMES:
            continue
        description = values.get("Visão Geral do Produto")
        area = values.get("Área")
        country = values.get("Região/País")
        platform = values.get("Plataformas")
        model = values.get("Modelo de relacionamento (SAAS/Venda/outros)")
        site = links.get(name, "")
        if not site:
            continue
        regions = infer_regions(country, site)
        yield {
            "name": name,
            "website": site,
            "scope": "competes_market" if "BR" in regions else "benchmark_global",
            "regions": regions,
            "markets": ["public", "private"],
            "tags": infer_tags(description, area, platform, model),
            "products_impacted": ["Aprimora"],
        }


def load_educacional_matrix():
    workbook = openpyxl.load_workbook(EDUCACIONAL_MATRIX, data_only=True, read_only=True)
    sheet = workbook["Competitors and key words"]
    current_label = ""
    for row in range(2, sheet.max_row + 1):
        label = clean(sheet.cell(row, 1).value)
        if label:
            current_label = label
        name = clean(sheet.cell(row, 2).value)
        site = split_site(sheet.cell(row, 3).value)
        if not name or not site or name in SKIP_NAMES:
            continue
        description = sheet.cell(row, 4).value
        differentiators = sheet.cell(row, 5).value
        audience = sheet.cell(row, 6).value
        products = infer_products(current_label, name)
        regions = infer_regions(site, description, audience)
        yield {
            "name": name,
            "website": site,
            "scope": "competes_market" if "BR" in regions else "benchmark_global",
            "regions": regions,
            "markets": ["public", "private"],
            "tags": infer_tags(current_label, name, description, differentiators, audience),
            "products_impacted": products,
        }


def write_yaml(competitors):
    lines = [
        'version: "0.1"',
        "notes:",
        '  - "Lista enriquecida com dados das planilhas de análise competitiva Aprimora e Educacional."',
        '  - "scope: competes_market (concorrência no nosso mercado) | benchmark_global (referência global)."',
        '  - "regions: BR/LATAM/GLOBAL. markets: public/private."',
        "competitors:",
    ]
    for item in sorted(competitors, key=lambda row: row["name"].lower()):
        lines.extend(
            [
                f"  - name: {quote(item['name'])}",
                f"    website: {quote(item.get('website', ''))}",
                f"    scope: {quote(item.get('scope', 'competes_market'))}",
                f"    regions: {flow(item.get('regions', ['BR']))}",
                f"    markets: {flow(item.get('markets', ['private']))}",
                f"    tags: {flow(item.get('tags', []))}",
                f"    products_impacted: {flow(item.get('products_impacted', []))}",
                "",
            ]
        )
    CONFIG_PATH.write_text("\n".join(lines), encoding="utf-8")


def main():
    merged = {}
    for item in read_current_competitors():
        upsert(merged, item)
    before = len(merged)
    for item in load_aprimora_matrix():
        upsert(merged, item)
    for item in load_educacional_matrix():
        upsert(merged, item)
    write_yaml(list(merged.values()))
    print(f"competitors_before={before}")
    print(f"competitors_after={len(merged)}")


if __name__ == "__main__":
    main()
